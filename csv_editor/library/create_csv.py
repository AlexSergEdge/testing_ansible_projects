#!/usr/bin/env python
#coding=utf-8
from ansible.module_utils.basic import *
from openpyxl import Workbook, cell
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font, PatternFill, Border, Protection, Alignment, Side
import os.path
import sys, os

# Функция используется в copy_header для того, что бы скоприровать полный стиль ячейки
def copyStyle(toStyle, fromStyle):
    toStyle.font = copy(fromStyle.font)
    toStyle.fill = copy(fromStyle.fill)
    toStyle.border = copy(fromStyle.border)
    toStyle.alignment = copy(fromStyle.alignment)
    toStyle.number_format = copy(fromStyle.number_format)
    toStyle.protection = copy(fromStyle.protection)

# Функция сравнения стилей двух ячеек
def areStylesEqual(toStyle, fromStyle):
    return (toStyle.font.name == fromStyle.font.name
            and toStyle.font.size == fromStyle.font.size
            and toStyle.font.bold == fromStyle.font.bold
            and toStyle.font.italic == fromStyle.font.italic
            and toStyle.font.vertAlign == fromStyle.font.vertAlign
            and toStyle.font.underline == fromStyle.font.underline
            and toStyle.font.strike == fromStyle.font.strike
            and toStyle.font.color == fromStyle.font.color
            
            and toStyle.fill.fill_type == fromStyle.fill.fill_type
            and toStyle.fill.start_color == fromStyle.fill.start_color
            and toStyle.fill.end_color == fromStyle.fill.end_color
            
            and toStyle.border.left.border_style == fromStyle.border.left.border_style
            and toStyle.border.left.color == fromStyle.border.left.color
            and toStyle.border.right.border_style == fromStyle.border.right.border_style
            and toStyle.border.right.color == fromStyle.border.right.color
            and toStyle.border.top.border_style == fromStyle.border.top.border_style
            and toStyle.border.top.color == fromStyle.border.top.color
            and toStyle.border.bottom.border_style == fromStyle.border.bottom.border_style
            and toStyle.border.bottom.color == fromStyle.border.bottom.color
            and toStyle.border.diagonal.border_style == fromStyle.border.diagonal.border_style
            and toStyle.border.diagonal.color == fromStyle.border.diagonal.color
            and toStyle.border.diagonal_direction == fromStyle.border.diagonal_direction
            and toStyle.border.outline == fromStyle.border.outline
                         
            and toStyle.alignment.horizontal == fromStyle.alignment.horizontal
            and toStyle.alignment.vertical == fromStyle.alignment.vertical
            and toStyle.alignment.text_rotation == fromStyle.alignment.text_rotation
            and toStyle.alignment.wrap_text == fromStyle.alignment.wrap_text
            and toStyle.alignment.shrink_to_fit == fromStyle.alignment.shrink_to_fit
            and toStyle.alignment.indent == fromStyle.alignment.indent
            
            and toStyle.number_format == fromStyle.number_format
            
            and toStyle.protection.locked == fromStyle.protection.locked
            and toStyle.protection.hidden == fromStyle.protection.hidden)

# Функция берет входной файл, копирует в массив ячейки заголовка, затем создает новый выходной файл (или берет существующий) и копирует ячейки в него
def copy_header(input_excel, output_excel):
    input_workbook = Workbook()
    output_workbook = Workbook()
    # загружаем пустую таблицу    
    input_workbook = load_workbook(input_excel)
    # выбираем рабочий лист
    input_worksheet = input_workbook["Sheet1"]
    # по умолчанию флаг изменения ставим в 0
    has_changed = False
    # объявляем массиво для данных заголовка
    header_values = []
    # Проходя по ячейкам сохраняем значения исходного заголовка
    for col in input_worksheet.iter_cols():
        for cell in col:
                header_values.append(cell.value)
    # Существует ли уже файл
    if (os.path.isfile(output_excel)):
        # Проверяем, соответсвует ли заголовок заголовку исходного файла 
        output_workbook = load_workbook(output_excel)
        output_worksheet = output_workbook["Sheet1"]
        for i in range(1, len(header_values)+1):
            outputCell = output_worksheet.cell(row = 1, column = i)
            inputCell = input_worksheet.cell(row = 1, column = i)
            # если что-то не так, корректируем выходной файл 
            if (outputCell.value != header_values[i - 1]) or not areStylesEqual(outputCell,inputCell):
                outputCell.value = header_values[i - 1]
                copyStyle(outputCell,inputCell)
                has_changed = True
        # Сохраняем изменения
        output_workbook.save(output_excel)
    # Если выходного файла нет - создаем его
    else:
        has_changed = True
        # создаем новый workbook и worksheet
        output_workbook.create_sheet("Sheet1", 0)
        output_worksheet = output_workbook["Sheet1"]
        # Копируем значения и стили ячеек из старого файла в новый
        for i in range(1, len(header_values)+1):
            outputCell = output_worksheet.cell(row = 1, column = i)
            inputCell = input_worksheet.cell(row = 1, column = i)
            outputCell.value = header_values[i - 1]
            copyStyle(outputCell,inputCell)
        # Сохраняем изменения
        output_workbook.save(output_excel)
    return False, has_changed, len(header_values)


# Функция копирует данные из массива в ячейки выходного файла
def copy_data(header_data, table_data, output_excel, real_header_length):

    output_workbook = Workbook()
    data_has_changed = False
    output_workbook = load_workbook(output_excel)
    output_worksheet = output_workbook["Sheet1"]
    # Проходим по ячейкам выходного файла и таблицчным данным
    for i in range(0, len(table_data[0])):
        for j in range(0, real_header_length):
            # В столбце, где заголовок соотвествует необходимому значению, заполняем все строки соотвествующими данными
            if header_data[i] == output_worksheet.cell(row=1, column=j+1).value:
                for k in range(0, len(table_data)):
                    # Проверка на существование или корректность существующих данных
                    if output_worksheet.cell(row=k+2, column=j+1).value != table_data[k][i]:
                        output_worksheet.cell(row=k+2, column=j+1).value = table_data[k][i]
                        data_has_changed = True
    # Сохраняем изменения
    output_workbook.save(output_excel)
    return False, data_has_changed


# Function saves input file in memory, fills it with data, and saves it as a new output file
# функция сохраняет входной файл, заполняет данными и создает новый выходной файл на его основе
# в отличие от метода 1 не требуется копировать заголовок
def create_csv(header_data, input_excel, output_excel, table_data):
    has_changed = True
    # Создаем workbook
    input_workbook = Workbook()
    # Открываем файл
    input_workbook = load_workbook(input_excel)
    # Выбираем worksheet
    input_worksheet = input_workbook["Sheet1"]
    # Реальный размер заголовка (в исходном файле)
    header_size=0
    for col in input_worksheet.iter_cols():
        for cell in col:
            if cell.value != '':
                header_size += 1
    # Заполняем ячейки данными
    for i in range(0, len(table_data[0])):
        for j in range(0, header_size):
            # если нашли нжный столбец, заполняем его строки данными
            if header_data[i] == input_worksheet.cell(row=1, column=j+1).value:
                for k in range(0, len(table_data)):
                    input_worksheet.cell(row=k+2, column=j+1).value = table_data[k][i]
    # Сохраняем выходной файл
    input_workbook.save(output_excel)
    return False, has_changed


def run_module():
    fields = dict(
        header_data = dict(required=True, type='list'),
        function_name = dict(default="full copy", choises=["full copy","workbook copy"], type='str'),
        table_data = dict(required=True, type='list'),
        input_excel = dict(required=True, type='str'),
        output_excel1 = dict(required=True, type='str'),
        output_excel2 = dict(required=True, type='str')   
    )
    
    result = dict(
        changed=False,
        message=''
    )
    
    module = AnsibleModule(argument_spec=fields, supports_check_mode=True)
    
    if module.check_mode:
        return result
    
    # входные данные получаем здесь
    header_data = module.params['header_data']
    function_name = module.params['function_name']
    table_data = module.params['table_data']
    input_excel = module.params['input_excel']
    output_excel1 = module.params['output_excel1']
    output_excel2 = module.params['output_excel2']
    
    try:
        # Если выбран первый способ
        if function_name == "full copy":
            # Копируем заголовок
            header_failed, header_changed, real_header_length = copy_header(input_excel, output_excel1)
            # Копируем данные
            data_failed, data_changed = copy_data(header_data, table_data, output_excel1, real_header_length)
            
            result['changed'] = data_changed or header_changed
            # если какие-либо данные поменялись
            if result['changed']:
                result['message'] = "Successfully copied excel data"
                module.exit_json(**result)
            # если файл уже был корректен
            else:
                result['message'] = "File already exists and correct. Do nothing"
                module.exit_json(**result)
        # Если выбран второй способ
        elif function_name == "workbook copy":
            # Вызываем функцию create_csv
            is_failed, is_changed = create_csv(header_data, input_excel, output_excel2, table_data)
            if is_changed:
                module.exit_json(changed=is_changed, failed=is_failed, meta="Successfully created new file")
            else:
                module.exit_json(changed=is_changed, failed=is_failed, meta="File already exists and correct. Do nothing")
    except IOError:
        module.exit_json(changed=False, failed=True, meta="File cannot be open")

def main():
    run_module()

if __name__ == '__main__':
    main()
