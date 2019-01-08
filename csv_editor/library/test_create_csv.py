#coding=utf-8
from __future__ import print_function
import pytest
import create_csv # импортируем модуль для тестирования
import json
import sys
import os
from ansible.module_utils._text import to_bytes
from ansible.module_utils import basic
from ansible.compat.tests import unittest
from cStringIO import StringIO
from openpyxl import Workbook, cell
from openpyxl.styles import Font, PatternFill, Border, Protection, Alignment, Side

# Далее приведены тестовые данные для модуля create_csv 

# Данные 1 (модуль должен создавать новые файлы при первом вызове и ничего не делать при повторном вызовах)
EXCEL_DATA_1 = {'header_data':['Name A', 'Interface A', 'Name B', 'Interface B'],
                    'function_name':'full copy',
                    'table_data':[['D1','E1','D2','E2'],['D3','E3','D4','E4'],['D5','E5','D6','E6'],],
                    'input_excel':'/syncfolder/excel_test_files/input_excel_uri.xlsx',
                    'output_excel1':'/syncfolder/excel_test_files/output_excel_uri1.xlsx',
                    'output_excel2':'/syncfolder/excel_test_files/output_excel_uri2.xlsx',
}

# Измененные данные 1 с аналогичными выходными файлами (модуль должен переписать файлы если был вызван после данных 1)
EXCEL_DATA_2 = {'header_data':['Name A', 'Interface A', 'Name B', 'Interface B'],
                    'function_name':'full copy',
                    'table_data':[['D100','E100','D200','E200'],['D300','E300','D400','E400'],['D500','E500','D600','E600'],],
                    'input_excel':'/syncfolder/excel_test_files/input_excel_uri.xlsx',
                    'output_excel1':'/syncfolder/excel_test_files/output_excel_uri1.xlsx',
                    'output_excel2':'/syncfolder/excel_test_files/output_excel_uri2.xlsx',
}

# Данные с измененными выходными файлами (при первом вызове создают новые файлы)
EXCEL_DATA_3 = {'header_data':['Name A', 'Interface A', 'Name B', 'Interface B'],
                    'function_name':'full copy',
                    'table_data':[['D100','E100','D200','E200'],['D300','E300','D400','E400'],['D500','E500','D600','E600'],],
                    'input_excel':'/syncfolder/excel_test_files/input_excel_uri.xlsx',
                    'output_excel1':'/syncfolder/excel_test_files/output_excel_uri1_new.xlsx',
                    'output_excel2':'/syncfolder/excel_test_files/output_excel_uri2_new.xlsx',
}

# Данные с неверным входным файлом - должна быть вызвана ошибка
NO_INPUT_FILE_EXCEL_DATA = {'header_data':['Name A', 'Interface A', 'Name B', 'Interface B'],
                    'function_name':'full copy',
                    'table_data':[['D1','E1','D2','E2'],['D3','E3','D4','E4'],['D5','E5','D6','E6'],],
                    'input_excel':'/syncfolder/no_excel_test_files/no_input_excel_uri.xlsx',
                    'output_excel1':'/syncfolder/excel_test_files/output_excel_uri1_new.xlsx',
                    'output_excel2':'/syncfolder/excel_test_files/output_excel_uri2_new.xlsx',
}

# Функия устанавливает аргументы на вход модуля
def set_module_args(args):
    args = json.dumps({"ANSIBLE_MODULE_ARGS": args})
    basic._ANSIBLE_ARGS = to_bytes(args)

# https://pytest.readthedocs.io/en/2.8.7/yieldfixture.html 
# функция задает необходимые действия при вызове теста несколько раз
# используется yield_fixture, которая сохраняет текущее состояние stdout и разрешает считывать данные из входного потока
@pytest.yield_fixture
def resource():
    backup = sys.stdout
    sys.stdout = StringIO()   
    yield "resource"   
    sys.stdout.close()
    sys.stdout = backup


# Класс обеспечивает тестирование на разных входных данных  
class TestExcelModule(object):  
    # параметризуем тесты
    @pytest.mark.parametrize('test_input,expected', [
        (EXCEL_DATA_1, 'Successfully copied excel data'),
        (EXCEL_DATA_1, 'File already exists and correct. Do nothing'),
        (EXCEL_DATA_2, 'Successfully copied excel data'),
        (EXCEL_DATA_3, 'Successfully copied excel data'),
        (NO_INPUT_FILE_EXCEL_DATA, 'File cannot be open'),
    ])    
    def test_create_csv_has_correct_output(self, test_input, expected, resource):
        # Для проверки происходит удаление файлов, если на вход поступили данные с пустым входным файлом
        # В дальнейшем это позволит повторно вызывать тесты 
        if (test_input == NO_INPUT_FILE_EXCEL_DATA):
            if os.path.isfile(EXCEL_DATA_1['output_excel1']):
                os.remove(EXCEL_DATA_1['output_excel1'])
            if os.path.isfile(EXCEL_DATA_3['output_excel1']):
                os.remove(EXCEL_DATA_3['output_excel1'])
        # Устанавливаем аргументы
        set_module_args(test_input)
        # вызываем основную функцию с выставленными на вход параметрами
        with pytest.raises(SystemExit):
            create_csv.main()
        # Получаем выходные данные и проверяем, что результат работы соотвествует ожиданиям
        output = sys.stdout.getvalue()
        assert expected in output

# Более простой отдельный тест проверяет верность работы функции, проверяющей стили ячеек
def test_styles_are_equal_function():  
    new_workbook = Workbook()
    new_workbook.create_sheet("Sheet1", 0)
    new_worksheet1 = new_workbook["Sheet1"]
    new_font = Font(name='Calibri',
                    size=11,
                    bold=False,
                    italic=True,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
    new_fill = PatternFill(fill_type=None,
                    start_color='FFFFFFFF',
                    end_color='FF000000')
    
    new_border = Border(left=Side(border_style=None,
                                  color='FF000000'),
                        right=Side(border_style=None,
                                   color='FF000000'),
                        top=Side(border_style=None,
                                 color='FF000000'),
                        bottom=Side(border_style=None,
                                    color='FF000000'),
                        diagonal=Side(border_style=None,
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=True,
                        vertical=Side(border_style=None,
                                      color='FF000000'),
                        horizontal=Side(border_style=None,
                                        color='FF000000')

                        )
    
    new_alignment=Alignment(horizontal='general',
                            vertical='bottom',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=False,
                            indent=0)
    
    new_number_format = 'General'
    
    new_protection = Protection(locked=True,
                                hidden=False)
    # cоздаем две ячейки с одинаковыми параметрами
    cell1 = new_worksheet1.cell(row = 1, column = 1)
    cell1.font = new_font
    cell1.fill = new_fill
    cell1.border = new_border
    cell1.protection = new_protection
    cell1.alignment = new_alignment
    cell1.number_format = new_number_format
    
    cell2 = new_worksheet1.cell(row = 1, column = 2)
    cell2.font = new_font
    cell2.fill = new_fill
    cell2.border = new_border
    cell2.protection = new_protection
    cell2.alignment = new_alignment
    cell2.number_format = new_number_format
    result = create_csv.areStylesEqual(cell1, cell2)
    assert result == True
